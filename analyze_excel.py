import openpyxl
import json
import traceback
import sys

sys.stdout.reconfigure(encoding='utf-8')

try:
    file_path = r'c:\Users\ujan\Desktop\RFP\Athlete-Performance\Athlete Performance Tracker.xlsx'
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    out = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=False)):
            if i > 30:
                break
            
            row_data = []
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith('='):
                    row_data.append(f"FORMULA: {val}")
                else:
                    row_data.append(str(val) if val is not None else None)
            
            if any(row_data):
                rows.append(row_data)
                
        out[sheet_name] = rows
        
    with open('c:\\Users\\ujan\\Desktop\\RFP\\Athlete-Performance\\excel_analysis_full.json', 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2)
        
    print("Successfully wrote analysis to excel_analysis_full.json")

except Exception as e:
    print(f"Error: {e}")
    traceback.print_exc()
